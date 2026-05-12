# src/server/session_recorder.py
# 재활 세션 기록 저장 - 영상(.avi), 음성(.wav), 분석결과(.xlsx)

import cv2
import wave
import openpyxl
import os
from datetime import datetime

SAVE_DIR = "data/sessions"

def get_filename(patient_name: str, ext: str) -> str:
    # 날짜_환자명 형식으로 파일명 생성
    date = datetime.now().strftime("%Y-%m-%d")
    os.makedirs(SAVE_DIR, exist_ok=True)
    return os.path.join(SAVE_DIR, f"{date}_{patient_name}.{ext}")

def save_video(patient_name: str, frames: list, fps: int = 20):
    # 영상 저장 (.avi)
    if not frames:
        return

    path = get_filename(patient_name, "avi")
    h, w, _ = frames[0].shape
    out = cv2.VideoWriter(path, cv2.VideoWriter_fourcc(*'XVID'), fps, (w, h))

    for frame in frames:
        out.write(frame)

    out.release()
    print(f"영상 저장 완료: {path}")

def save_audio(patient_name: str, audio_data: bytes, sample_rate: int = 44100):
    # 음성 저장 (.wav)
    path = get_filename(patient_name, "wav")

    with wave.open(path, 'wb') as f:
        f.setnchannels(1)
        f.setsampwidth(2)  # int16 = 2바이트
        f.setframerate(sample_rate)
        f.writeframes(audio_data)

    print(f"음성 저장 완료: {path}")

def save_excel(patient_name: str, session_data: dict):
    # 분석 결과 저장 (.xlsx)
    # session_data 예시:
    # {"비대칭 지수": 0.12, "발음 정확도": 87, "발화 속도": 3.2, "음량": 0.8, "묵음 구간": 1.5}

    path = get_filename(patient_name, "xlsx")

    # 파일 있으면 불러오고 없으면 새로 생성
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["회차", "날짜", "비대칭 지수", "발음 정확도", "발화 속도", "음량", "묵음 구간"])

    row = ws.max_row
    ws.append([
        row,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        session_data.get("비대칭 지수", "-"),
        session_data.get("발음 정확도", "-"),
        session_data.get("발화 속도", "-"),
        session_data.get("음량", "-"),
        session_data.get("묵음 구간", "-"),
    ])

    wb.save(path)
    print(f"분석 결과 저장 완료: {path}")