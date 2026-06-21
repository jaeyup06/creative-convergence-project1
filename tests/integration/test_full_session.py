# tests/integration/test_full_session.py
# 세션 종료 파이프라인의 데이터 흐름 검증:
#   누적음성 -> analyze_voice -> session_data(한글키) -> save_excel(컬럼)
# server._finalize_session 과 동일한 키 매핑을 사용해, 컬럼 불일치로
# 값이 "-"로 새는 일이 없는지 확인한다. (TTS/네트워크 없이 동작)
import os
import numpy as np
import openpyxl

from src.recognition.voice_analyzer import analyze_voice
from src.server.session_recorder import save_excel, SAVE_DIR
from src.common.config import AUDIO_SAMPLE_RATE


def _make_buffer():
    sr = AUDIO_SAMPLE_RATE
    t = np.linspace(0, 2.0, int(sr * 2.0), endpoint=False)
    tone = (np.sin(2 * np.pi * 220 * t) * 8000).astype(np.int16)
    silence = np.zeros(int(sr * 1.0), dtype=np.int16)
    return np.concatenate([tone, silence]).tobytes()


def test_session_pipeline_writes_real_values():
    name = "테스트환자_pipeline"
    audio_bytes = _make_buffer()
    audio_np = np.frombuffer(audio_bytes, dtype=np.int16)

    voice = analyze_voice(audio_np, AUDIO_SAMPLE_RATE)
    session_data = {
        "비대칭 지수": "-",
        "발음 정확도": "-",
        "발화 속도": voice.get("speech_rate") if voice.get("speech_rate") is not None else "-",
        "음량": voice.get("volume") if voice.get("volume") is not None else "-",
        "묵음 구간": voice.get("silence_sec") if voice.get("silence_sec") is not None else "-",
    }
    save_excel(name, session_data)

    from datetime import datetime
    date = datetime.now().strftime("%Y-%m-%d")
    path = os.path.join(SAVE_DIR, f"{date}_{name}.xlsx")
    assert os.path.exists(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["회차", "날짜", "비대칭 지수", "발음 정확도", "발화 속도", "음량", "묵음 구간"]

    last = [c.value for c in ws[ws.max_row]]
    assert isinstance(last[4], (int, float))  # 발화 속도
    assert isinstance(last[5], (int, float))  # 음량
    assert isinstance(last[6], (int, float))  # 묵음 구간

    os.remove(path)


if __name__ == "__main__":
    test_session_pipeline_writes_real_values()
    print("full_session 파이프라인 테스트 통과")